# -*- coding: utf-8 -*-
"""
Created on Wed May  8 17:15:35 2019

@author: vpm6
"""

import pandas as pd
#import matplotlib.pyplot as plt
#import seaborn as sns
#import numpy as np

from openpyxl import load_workbook


#data = load_workbook(r'C:\Users\vpm6\Desktop\python\numpy\issue.xlsx')
data = load_workbook(r'C:\Users\vpm6\Desktop\python\numpy\issue.xlsx')
sheet_1=pd.read_excel(r'D:\python\numpy\issue.xslx',sheetname=0)
sheet_2=pd.read_excel(r'D:\python\numpy\issue.xslx',sheetname=1)
sheet_3=pd.read_excel(r'D:\python\numpy\issue.xslx',sheetname=2)


read_sheets_name = data.sheet_names
print (read_sheets_name)
